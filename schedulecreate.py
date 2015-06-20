__author__ = 'n3tn0'
__version__ = 'BETA'

class CreateSchedule():
    def getnames(self):
        # Get a list of all the workers with no duplicates
        workers = []
        for i in range(2, rows + 2):
            names = sheet_ranges['K'+str(i)].value
            for i in names.split(', '):
                if i not in workers:
                    workers.append(i)
                    workers.append('')
        return workers

    def getdates(self):
        from datetime import timedelta
        corenames = {}
        coredates = {}
        eventid = 0
        for i in range(2, rows + 2):
            startdate = sheet_ranges['G' + str(i)].value
            enddate = sheet_ranges['H' + str(i)].value

            # Take each date range and apply all of the dates in it to a dictionary object with the name of the worker
            curr = startdate
            while curr <= enddate:
                for j in sheet_ranges['K' + str(i)].value.split(', '):
                    corenames[eventid] = j
                    coredates[eventid] = curr.date()
                    eventid += 1
                curr += timedelta(days=1)
        return corenames, coredates

    def getextremes(self, core):
        import datetime
        now = datetime.datetime.now()

        lowest = now.today().date()
        for i in core:
            if core[i] < lowest:
                lowest = core[i]

        highest = now.today().date()
        for i in core:
            if core[i] > highest:
                highest = core[i]

        return lowest, highest

    def createcalendar(self, calendarfile, workers, corenames, coredates, lowestdate, highestdate):
        from dateutil.relativedelta import relativedelta
        import calendar
        import xlsxwriter
        calendarwb = xlsxwriter.Workbook(calendarfile)

        months = []
        curr = lowestdate
        while curr <= highestdate:
            months.append(curr)
            curr += relativedelta(months=1)

        # Set formatting rules
        headings = calendarwb.add_format()
        headings.set_bold()
        headings.set_border(1)

        fill = calendarwb.add_format()
        fill.set_bg_color('#118ED5')
        fill.set_border(1)
        #fill.set_width(256 * 3)

        emptyrow = calendarwb.add_format()
        emptyrow.set_bg_color('#808080')
        emptyrow.set_border()

        for i in months:
            workingwith = "%s %s" % (calendar.month_name[i.month], str(i.year))
            worksheet = calendarwb.add_worksheet(workingwith)

            # Write month and date to A1
            worksheet.write(0, 0, workingwith, headings)

            # Write day numbers
            days = calendar.monthrange(i.year, i.month)
            row = 0
            col = 1
            for j in range(1, days[1]):
                worksheet.write(row, col, j, headings)
                col += 1

            # Write list of names to column A
            row = 1
            col = 0
            k = 0
            while row < len(workers):
                worksheet.write(row, col, workers[k], headings)
                row += 1
                k += 1

            # Fill in empty rows
            m = 1
            for j in workers:
                if not j:
                    k = 0
                    while k < days[1]:
                        worksheet.write(workers.index(j) + m, k, '', emptyrow)
                        k += 1
                    m += 2

            # Fill in the chart!
            for j in coredates:
                x = 1
                if coredates[j].month == i.month:
                    worksheet.write(workers.index(corenames[j]) + x, coredates[j].day, '', fill)
                    x += 2

    def build(self, workbook, calendarfile):
        # Setup workbook for usage
        from openpyxl import load_workbook
        wb = load_workbook(workbook)
        global sheet, sheet_ranges, rows
        sheet = wb.worksheets[0]
        sheet_ranges = wb['Sheet1']
        rows = sheet.get_highest_row() - 1

        workers = self.getnames()
        corenames, coredates = self.getdates()
        lowestdate, highestdate = self.getextremes(coredates)
        self.createcalendar(calendarfile, workers, corenames, coredates, lowestdate, highestdate)